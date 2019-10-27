from django.shortcuts import render, redirect
import cv2
import numpy as np


from . import cascade as casc
from PIL import Image
import os
import face_recognition


import sys



import pickle

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Create your views here.
def index(request):
    return render(request, 'index.html')
def errorImg(request):
    return render(request, 'error.html')

def create_dataset(request):
    #print request.POST
    userId = request.POST['userId']
    
    # Detect face
    #Creating a cascade image classifier
    faceDetect = cv2.CascadeClassifier(BASE_DIR+'/ml/haarcascade_frontalface_default.xml')
    #camture images from the webcam and process and detect the face
    # takes video capture id, for webcam most of the time its 0.
    cam = cv2.VideoCapture(0)

    # Our identifier
    # We will put the id here and we will store the id with a face, so that later we can identify whose face it is
    id = userId
    # Our dataset naming counter
    sampleNum = 0
    # Capturing the faces one by one and detect the faces and showing it on the window
    while(True):
        # Capturing the image
        #cam.read will return the status variable and the captured colored image
        ret, img = cam.read()
        #the returned img is a colored image but for the classifier to work we need a greyscale image
        #to convert
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        #To store the faces
        #This will detect all the images in the current frame, and it will return the coordinates of the faces
        #Takes in image and some other parameter for accurate result
        faces = faceDetect.detectMultiScale(gray, 1.3, 5)
        #In above 'faces' variable there can be multiple faces so we have to get each and every face and draw a rectangle around it.
        for(x,y,w,h) in faces:
            # Whenever the program captures the face, we will write that is a folder
            # Before capturing the face, we need to tell the script whose face it is
            # For that we will need an identifier, here we call it id
            # So now we captured a face, we need to write it in a file
            sampleNum = sampleNum+1
            # Saving the image dataset, but only the face part, cropping the rest
            cv2.imwrite(BASE_DIR+'/ml/dataset/user.'+str(id)+'.'+str(sampleNum)+'.jpg', gray[y:y+h,x:x+w])
            # @params the initial point of the rectangle will be x,y and
            # @params end point will be x+width and y+height
            # @params along with color of the rectangle
            # @params thickness of the rectangle
            cv2.rectangle(img,(x,y),(x+w,y+h), (0,255,0), 2)
            # Before continuing to the next loop, I want to give it a little pause
            # waitKey of 100 millisecond
            cv2.waitKey(250)

        #Showing the image in another window
        #Creates a window with window name "Face" and with the image img
        cv2.imshow("Face",img)
        #Before closing it we need to give a wait command, otherwise the open cv wont work
        # @params with the millisecond of delay 1
        cv2.waitKey(1)
        #To get out of the loop
        if(sampleNum>35):
            break
    #releasing the cam
    cam.release()
    # destroying all the windows
    cv2.destroyAllWindows()

    return redirect('/')

def trainer(request):
    '''
        In trainer.py we have to get all the samples from the dataset folder,
        for the trainer to recognize which id number is for which face.

        for that we need to extract all the relative path
        i.e. dataset/user.1.1.jpg, dataset/user.1.2.jpg, dataset/user.1.3.jpg
        for this python has a library called os
    '''
    import face_recognition
    import os
    from face_recognition.face_recognition_cli import image_files_in_folder
    import sys
    import cv2


    path = BASE_DIR+'/ml/dataset'

# To get all the images, we need corresponing id

# create a list for the path for all the images that is available in the folder
# from the path(dataset folder) this is listing all the directories and it is fetching the directories from each and every pictures
# And putting them in 'f' and join method is appending the f(file name) to the path with the '/'
    imagePaths = [os.path.join(path,f) for f in os.listdir(path)] #concatinate the path with the image name
#print imagePaths

# Now, we loop all the images and store that userid and the face with different image list
    faces = []
    Ids = []
    for imagePath in imagePaths:
	    # First we have to open the image then we have to convert it into numpy array
	    faceImg = face_recognition.load_image_file(imagePath) #convert it to grayscale
	    # converting the PIL image to numpy array
	    # @params takes image and convertion format
	    faceNp = face_recognition.face_encodings(faceImg)[0]
	    # Now we need to get the user id, which we can get from the name of the picture
	    # for this we have to slit the path() i.e dataset/user.1.7.jpg with path splitter and then get the second part only i.e. user.1.7.jpg
	    # Then we split the second part with . splitter
	    # Initially in string format so hance have to convert into int format
	    ID = int(os.path.split(imagePath)[-1].split('.')[1]) # -1 so that it will count from backwards and slipt the second index of the '.' Hence id
	    # Images
	    faces.append(faceNp)
	    # Label
	    Ids.append(ID)
	    #print ID
	    #cv2.imshow("training", faceNp)
	    #cv2.waitKey(10)
	    
	
	# Fetching ids and faces
    data = {"encodings": faces, "names": Ids}


    


    
    encoding_faces = 'D:/Download/faceRecog-20190421T090705Z-001/faceRecog/encoding_faces.pkl'
    f = open(encoding_faces, "wb")
    f.write(pickle.dumps(data))
    f.close()

    return redirect('/')


def detect(request):
    video_capture = cv2.VideoCapture(0)
    encod_face = 'D:/Download/faceRecog-20190421T090705Z-001/faceRecog/encoding_faces.pkl'

    x = open( encod_face, "rb")
    data = pickle.load(x)
    while True:
    # Grab a single frame of video
        ret, frame = video_capture.read()

        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        rgb_frame = frame[:, :, ::-1]

        # Find all the faces and face enqcodings in the frame of video
        face_locations = face_recognition.face_locations(rgb_frame,number_of_times_to_upsample=2)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        face_names = []
        # Loop through each face in this frame of video
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
			
        # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(data["encodings"], face_encoding,tolerance=0.45)

            name = "Unknown"

        # If a match was found in known_face_encodings, just use the first one.
        # if True in matches:
        #     first_match_index = matches.index(True)
        #     name = known_face_names[first_match_index]

        # Or instead, use the known face with the smallest distance to the new face
            face_distances = face_recognition.face_distance(data["encodings"], face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = data["names"][best_match_index]
                face_names.append(name)
			
			
			
				
            

        
            
            
            
			      

        # Display the resulting image
        cv2.imshow('Video', frame)

        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release handle to the webcam
    video_capture.release()
    
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'D:/Download/faceRecog-20190421T090705Z-001/faceRecog/Attendance_List.xlsx')
    ws = wb['Sheet']
    import pandas as pd
    date = pd.datetime.now().date()
    time = pd.datetime.now().time()
    ws.cell(row = 1, column = 4, value = date)
    for i in range(2,ws.max_row+1):
        ws.cell(row=i, column=4).value="Absent"
    for name in face_names:
        for i in range(2,ws.max_row+1):
            if(name==((ws.cell(row=i, column=2).value))):
                ws.cell(row=i, column=4).value = "Present"
                ws.cell(row=i , column = 5).value = time
            
        
    wb.save("D:/Download/faceRecog-20190421T090705Z-001/faceRecog/Attendance_List.xlsx")    
        
       
    print("Students present are as follows: ")   
    for name in face_names:
         print(name)
    
    
    cv2.destroyAllWindows()
    return redirect('/')



def detectImage(request):
    faceDetect = cv2.CascadeClassifier(BASE_DIR+'/ml/haarcascade_frontalface_default.xml')

    cam = cv2.VideoCapture(0)
    # creating recognizer
    rec = cv2.face.LBPHFaceRecognizer_create();
    # loading the training data
    rec.read(BASE_DIR+'/ml/recognizer/trainingData.yml')
    getId = 0
    font = cv2.FONT_HERSHEY_SIMPLEX
    userId = 0
    while(True):
        ret, img = cam.read()
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = faceDetect.detectMultiScale(gray, 1.3, 5)
        for(x,y,w,h) in faces:
            cv2.rectangle(img,(x,y),(x+w,y+h), (0,255,0), 2)

            getId,conf = rec.predict(gray[y:y+h, x:x+w]) #This will predict the id of the face

            #print conf;
            if conf<35:
                userId = getId
                name = str(userId)
                cv2.putText(img, name,(x,y+h), font, 2, (0,255,0),2)
            else:
                cv2.putText(img, "Unknown",(x,y+h), font, 2, (0,0,255),2)

            # Printing that number below the face
            # @Prams cam image, id, location,font style, color, stroke

        cv2.imshow("Face",img)
        if(cv2.waitKey(1) == ord('q')):
            break
        elif(userId != 0):
            cv2.waitKey(1000)
            cam.release()
            cv2.destroyAllWindows()
            return redirect('/records/details/'+str(userId))

    cam.release()
    cv2.destroyAllWindows()
    return redirect('/')

def eigenTrain(request):
    import webbrowser
    webbrowser.open('D:/Download/faceRecog-20190421T090705Z-001/faceRecog/Attendance_List.xlsx')
    return redirect('/')


